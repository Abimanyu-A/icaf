from pathlib import Path
import yaml
from openpyxl import load_workbook


class ProfileLoader:
    def __init__(self, profile_name):
        profile_dir = Path("icaf/profile")

        yaml_path = profile_dir / f"{profile_name}.yaml"
        yml_path = profile_dir / f"{profile_name}.yml"
        xlsx_path = profile_dir / f"{profile_name}.xlsx"

        if yaml_path.exists():
            self.data = self._load_yaml(yaml_path)
        elif yml_path.exists():
            self.data = self._load_yaml(yml_path)
        elif xlsx_path.exists():
            self.data = self._load_xlsx(xlsx_path)
        else:
            raise FileNotFoundError(
                f"Profile '{profile_name}' not found. "
                f"Expected one of: {yaml_path.name}, {yml_path.name}, {xlsx_path.name}"
            )

    def _load_yaml(self, path):
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}

    def _load_xlsx(self, path):
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        data = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            key, value = row[:2]

            if key is None:
                continue

            key = str(key).strip()
            value = "" if value is None else str(value).strip()

            self._insert(data, key, value)

        return data

    def _insert(self, data, dotted_key, value):
        keys = dotted_key.split(".")
        node = data

        for k in keys[:-1]:
            if k not in node or not isinstance(node[k], dict):
                node[k] = {}
            node = node[k]

        last_key = keys[-1]

        if last_key in node:
            if isinstance(node[last_key], list):
                node[last_key].append(value)
            else:
                node[last_key] = [node[last_key], value]
        else:
            node[last_key] = value

    def get(self, key, default=None):
        keys = key.split(".")
        value = self.data

        for k in keys:
            if not isinstance(value, dict) or k not in value:
                return default
            value = value[k]

        return value

    def get_list(self, key, default=None):
        value = self.get(key, default if default is not None else [])
        if value is None:
            return []
        if isinstance(value, list):
            return value
        return [value]