"""
OEM Excel reader - ported from 2.6.1/OEM_TEST_CASES/OEM_TC1.py
Reads the OEM-supplied protocol declaration spreadsheet.
"""

import os
from openpyxl import load_workbook


def read_oem_data(oem_file: str) -> dict:
    """
    Parse an OEM protocol declaration Excel file.

    Expected columns (no header row required):
        col A – protocol name
        col B – supported (Yes/No or similar)
        col C – details / remarks

    Returns:
        dict mapping protocol name → {"supported": str, "details": str}
    """
    workbook = load_workbook(oem_file)
    sheet = workbook.active

    protocol_data = {}

    for row in sheet.iter_rows(values_only=True):
        if not row or not row[0]:
            continue  # skip empty rows

        protocol = str(row[0]).strip()
        supported = str(row[1]).strip() if row[1] else ""
        details = str(row[2]).strip() if row[2] else ""

        protocol_data[protocol] = {
            "supported": supported,
            "details": details,
        }

    return protocol_data


def run_oem_test(oem_file: str | None = None) -> dict:
    """
    Run the OEM declaration test.

    Args:
        oem_file: Path to the OEM Excel file.  If None the function looks
                  for ``OEM.xlsx`` next to this module.

    Returns:
        {"oem_protocol_data": dict}
    """
    if oem_file is None:
        oem_file = os.path.join(os.path.dirname(__file__), "OEM.xlsx")

    return {"oem_protocol_data": read_oem_data(oem_file)}
